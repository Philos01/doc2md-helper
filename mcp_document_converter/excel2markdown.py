import os
import sys
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

DEFAULT_TIMEOUT = 300


def excel_to_markdown(excel_path: str, output_dir: str = None) -> str:
    """
    将Excel文件转换为Markdown格式，支持 .xlsx 和 .xls 格式

    Args:
        excel_path: Excel文件路径
        output_dir: 输出目录，如果为None则使用Excel文件所在目录

    Returns:
        转换后的Markdown文件路径

    Raises:
        Exception: 转换失败时抛出异常
    """
    if output_dir is None:
        output_dir = os.path.dirname(excel_path)

    os.makedirs(output_dir, exist_ok=True)

    timeout = DEFAULT_TIMEOUT

    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.md")
    file_ext = os.path.splitext(excel_path)[1].lower()

    if file_ext not in ['.xlsx', '.xls']:
        raise ValueError(f"不支持的文件格式: {file_ext}，仅支持 .xlsx 和 .xls")

    try:
        # 方法1: 使用 MarkItDown (推荐)
        try:
            from markitdown import MarkItDown

            logger.info(f"[DEBUG] 使用MarkItDown转换Excel文件: {excel_path}")

            md = MarkItDown()
            result = md.convert(excel_path)
            markdown_text = result.text_content

            # 保存Markdown文件
            with open(output_path, "w", encoding="utf-8") as md_file:
                md_file.write(markdown_text)

            logger.info(f"[DEBUG] MarkItDown转换成功，Markdown文件: {output_path}")
            return output_path

        except ImportError:
            logger.warning("[DEBUG] MarkItDown库未安装，尝试备选方案")
        except Exception as e:
            logger.warning(f"[DEBUG] MarkItDown转换失败: {e}，尝试备选方案")

        # 方法2: 使用 openpyxl + 手动格式化 (适用于 .xlsx)
        if file_ext == '.xlsx':
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter

                logger.info(f"[DEBUG] 使用openpyxl转换Excel文件: {excel_path}")

                wb = openpyxl.load_workbook(excel_path, data_only=True)
                markdown_lines = []

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    # 添加工作表标题
                    markdown_lines.append(f"## {sheet_name}")
                    markdown_lines.append("")

                    # 处理每个单元格
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        # 跳过空行
                        if all(cell is None for cell in row):
                            continue

                        # 格式化行数据
                        cells = []
                        for cell in row:
                            if cell is None:
                                cells.append("")
                            elif isinstance(cell, (int, float)):
                                cells.append(str(cell))
                            else:
                                # 转义特殊字符
                                cell_str = str(cell)
                                cell_str = cell_str.replace('|', '\\|')
                                cell_str = cell_str.replace('\n', ' ')
                                cells.append(cell_str.strip())

                        # 添加表格行
                        markdown_lines.append("| " + " | ".join(cells) + " |")

                        # 添加表头分隔符（仅第一行）
                        if row_idx == 1:
                            markdown_lines.append("| " + " | ".join(["---"] * len(cells)) + " |")

                    markdown_lines.append("")

                # 处理合并单元格（如果存在）
                if sheet.merged_cells:
                    logger.info(f"[DEBUG] 检测到合并单元格，将尽量保留结构")

                markdown_text = "\n".join(markdown_lines)

                # 保存Markdown文件
                with open(output_path, "w", encoding="utf-8") as md_file:
                    md_file.write(markdown_text)

                logger.info(f"[DEBUG] openpyxl转换成功，Markdown文件: {output_path}")
                return output_path

            except ImportError:
                logger.error("[DEBUG] openpyxl库未安装")
            except Exception as e:
                logger.error(f"[DEBUG] openpyxl转换失败: {e}")

        # 方法3: 使用 pandas + 格式化输出
        try:
            import pandas as pd

            logger.info(f"[DEBUG] 使用pandas转换Excel文件: {excel_path}")

            # 读取所有工作表
            excel_file = pd.ExcelFile(excel_path)
            markdown_lines = []

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # 添加工作表标题
                markdown_lines.append(f"## {sheet_name}")
                markdown_lines.append("")

                # 生成表头
                headers = df.columns.tolist()
                markdown_lines.append("| " + " | ".join(str(h) for h in headers) + " |")
                markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

                # 添加数据行
                for _, row in df.iterrows():
                    cells = []
                    for val in row:
                        if pd.isna(val):
                            cells.append("")
                        elif isinstance(val, (int, float)):
                            cells.append(str(val))
                        else:
                            cell_str = str(val)
                            cell_str = cell_str.replace('|', '\\|')
                            cells.append(cell_str.strip())
                    markdown_lines.append("| " + " | ".join(cells) + " |")

                markdown_lines.append("")

            markdown_text = "\n".join(markdown_lines)

            # 保存Markdown文件
            with open(output_path, "w", encoding="utf-8") as md_file:
                md_file.write(markdown_text)

            logger.info(f"[DEBUG] pandas转换成功，Markdown文件: {output_path}")
            return output_path

        except ImportError:
            logger.error("[DEBUG] pandas库未安装")
        except Exception as e:
            logger.error(f"[DEBUG] pandas转换失败: {e}")

        # 所有方法都失败了
        raise Exception(
            f"Excel文件转换失败，已尝试的方法:\n"
            f"  1. MarkItDown (推荐) - 需要安装: pip install 'markitdown[xlsx]'\n"
            f"  2. openpyxl - 需要安装: pip install openpyxl\n"
            f"  3. pandas - 需要安装: pip install pandas openpyxl\n"
            f"请至少安装其中一个库"
        )

    except Exception as e:
        logger.error(f"[ERROR] Excel转换失败: {e}")
        import traceback
        traceback.print_exc()
        raise


def excel_to_markdown_simple(excel_path: str) -> str:
    """
    简单的Excel转Markdown转换，直接返回Markdown文本，不保存文件

    Args:
        excel_path: Excel文件路径

    Returns:
        Markdown格式的字符串
    """
    file_ext = os.path.splitext(excel_path)[1].lower()

    if file_ext not in ['.xlsx', '.xls']:
        raise ValueError(f"不支持的文件格式: {file_ext}")

    try:
        # 尝试使用 MarkItDown
        try:
            from markitdown import MarkItDown
            md = MarkItDown()
            result = md.convert(excel_path)
            return result.text_content
        except (ImportError, Exception):
            pass

        # 备选：使用 openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            markdown_lines = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                markdown_lines.append(f"## {sheet_name}\n")

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if all(cell is None for cell in row):
                        continue

                    cells = []
                    for cell in row:
                        if cell is None:
                            cells.append("")
                        elif isinstance(cell, (int, float)):
                            cells.append(str(cell))
                        else:
                            cell_str = str(cell)
                            cell_str = cell_str.replace('|', '\\|')
                            cells.append(cell_str.strip())

                    markdown_lines.append("| " + " | ".join(cells) + " |")
                    if row_idx == 1:
                        markdown_lines.append("| " + " | ".join(["---"] * len(cells)) + " |")

                markdown_lines.append("")
            return "\n".join(markdown_lines)

        except ImportError:
            raise Exception("请安装 openpyxl 或 markitdown 库以支持Excel转换")

    except Exception as e:
        logger.error(f"[ERROR] Excel转Markdown失败: {e}")
        raise


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python excel2markdown.py <excel文件路径> [输出目录]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        output_path = excel_to_markdown(excel_path, output_dir)
        print(f"转换成功: {output_path}")
    except Exception as e:
        print(f"转换失败: {e}")
        sys.exit(1)
